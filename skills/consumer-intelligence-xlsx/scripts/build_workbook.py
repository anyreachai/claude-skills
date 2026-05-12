"""
Consumer Intelligence XLSX builder.

Reads a JSON config (see references/workbook-spec.md for schema) and produces
the standard 5-sheet workbook:
  1. Review Summary
  2. Third-Party Reviews
  3. Employee Reviews
  4. Competitive Benchmarking
  5. Competitor Deep Dives

Usage:
  python build_workbook.py --config <path.json> --output <path.xlsx>
"""

import argparse
import json
from collections import OrderedDict, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ---------------------------------------------------------------------------
# Design tokens
# ---------------------------------------------------------------------------

DARK_BLUE_FILL = 'FF1F3864'
MED_BLUE_FILL = 'FF2F5496'
LIGHT_BLUE_FILL = 'FFD6E4F0'

TITLE_FONT = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
SECTION_FONT = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
HEADER_FONT = Font(name='Calibri', size=11, bold=True)
NORMAL_FONT = Font(name='Calibri', size=11)
BOLD_FONT = Font(name='Calibri', size=11, bold=True)
NEG_FONT = Font(name='Calibri', size=11, bold=True, color='CC0000')
POS_FONT = Font(name='Calibri', size=11, bold=True, color='006600')
MIX_FONT = Font(name='Calibri', size=11, bold=True, color='CC6600')

THIN_BORDER = Border(bottom=Side(style='thin', color='D0D0D0'))


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def style_title(ws, row, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    cell.font = TITLE_FONT
    cell.fill = PatternFill('solid', fgColor=DARK_BLUE_FILL)
    cell.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 28


def style_section(ws, row, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    cell.font = SECTION_FONT
    cell.fill = PatternFill('solid', fgColor=MED_BLUE_FILL)
    cell.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 22


def style_headers(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = PatternFill('solid', fgColor=LIGHT_BLUE_FILL)
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )


def write_data_row(ws, row, data_list, font_overrides=None):
    """Write a data row. font_overrides is a list aligned with data_list; None means use NORMAL_FONT."""
    for i, val in enumerate(data_list):
        cell = ws.cell(row=row, column=i + 1, value=val)
        if font_overrides and i < len(font_overrides) and font_overrides[i] is not None:
            cell.font = font_overrides[i]
        else:
            cell.font = NORMAL_FONT
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = THIN_BORDER


def font_for_sentiment(value):
    if value == 'Negative':
        return NEG_FONT
    if value == 'Positive':
        return POS_FONT
    if value == 'Mixed':
        return MIX_FONT
    return None


def set_column_widths(ws, widths):
    """widths is a dict like {'A': 35, 'B': 18, ...}"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ---------------------------------------------------------------------------
# Platform grouping
# ---------------------------------------------------------------------------

def group_reviews_by_platform(reviews):
    """Group reviews by platform, ordered by review count descending."""
    grouped = defaultdict(list)
    for r in reviews:
        grouped[r['platform']].append(r)
    # Sort platforms by review count descending
    return OrderedDict(
        sorted(grouped.items(), key=lambda kv: -len(kv[1]))
    )


# ---------------------------------------------------------------------------
# Sheet 1: Review Summary
# ---------------------------------------------------------------------------

def build_review_summary(ws, cfg):
    set_column_widths(ws, {'A': 35, 'B': 18, 'C': 20, 'D': 55})
    max_col = 4

    company = cfg['company_name']
    prepared_by = cfg['prepared_by']
    focus_area = cfg.get('focus_area')
    focus_desc = cfg.get('focus_area_description', '')

    # Row 1 — Title
    if focus_area:
        title_text = f"{company} — {focus_area} Review Intelligence Report"
    else:
        title_text = f"{company} Review Intelligence Report"
    ws.cell(row=1, column=1, value=title_text)
    style_title(ws, 1, max_col)

    # Row 2 — Subtitle
    subtitle_parts = [f"Prepared by: {prepared_by}"]
    if focus_desc:
        subtitle_parts.append(focus_desc)
    ws.cell(row=2, column=1, value=" | ".join(subtitle_parts))
    style_section(ws, 2, max_col)

    row = 3

    # --- Section: Review Ratings by Platform ---
    ws.cell(row=row, column=1, value="Review Ratings by Platform")
    style_section(ws, row, max_col)
    row += 1

    headers = ["Platform", "Rating", "# of Reviews", "Notes"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    style_headers(ws, row, len(headers))
    row += 1

    for p in cfg.get('platforms', []):
        write_data_row(
            ws, row,
            [p.get('name'), p.get('avg_rating'), p.get('review_count'), p.get('notes', '')]
        )
        row += 1

    row += 1  # gap

    # --- Section: Key Themes Analysis ---
    ws.cell(row=row, column=1, value="Key Themes Analysis")
    style_section(ws, row, max_col)
    row += 1

    headers = ["Theme", "Sentiment", "Frequency %", "Confidence"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    style_headers(ws, row, len(headers))
    row += 1

    # Sort: negatives by freq desc, then positives, then mixed
    themes = list(cfg.get('themes', []))

    def theme_sort_key(t):
        sentiment_order = {'Negative': 0, 'Positive': 1, 'Mixed': 2}
        return (sentiment_order.get(t.get('sentiment'), 3), -t.get('frequency_pct', 0))

    themes.sort(key=theme_sort_key)

    for t in themes:
        sentiment_val = t.get('sentiment', '')
        font_overrides = [None, font_for_sentiment(sentiment_val), None, None]
        write_data_row(
            ws, row,
            [t.get('name'), sentiment_val, t.get('frequency_pct'), t.get('confidence', '')],
            font_overrides=font_overrides,
        )
        row += 1

    row += 1  # gap

    # --- Section: Sentiment Over Time ---
    ws.cell(row=row, column=1, value="Sentiment Over Time")
    style_section(ws, row, max_col)
    row += 1

    headers = ["Period", "Reviews", "Positive", "Negative"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    style_headers(ws, row, len(headers))
    row += 1

    for q in cfg.get('sentiment_timeline', []):
        font_overrides = [None, None, POS_FONT, NEG_FONT]
        write_data_row(
            ws, row,
            [q.get('period'), q.get('total'), q.get('positive'), q.get('negative')],
            font_overrides=font_overrides,
        )
        row += 1

    # --- Optional: Competitor Comparison ---
    comp_block = cfg.get('competitor_comparison_in_summary')
    if comp_block and comp_block.get('rows'):
        row += 1
        label = (
            f"Competitor {focus_area} Comparison" if focus_area
            else "Competitor Comparison"
        )
        ws.cell(row=row, column=1, value=label)
        style_section(ws, row, max_col)
        row += 1

        headers = comp_block.get('headers', [])
        for i, h in enumerate(headers):
            ws.cell(row=row, column=i + 1, value=h)
        style_headers(ws, row, len(headers))
        row += 1

        for data_row in comp_block.get('rows', []):
            write_data_row(ws, row, data_row)
            row += 1


# ---------------------------------------------------------------------------
# Sheet 2: Third-Party Reviews
# ---------------------------------------------------------------------------

def build_third_party_reviews(ws, cfg):
    set_column_widths(ws, {
        'A': 5, 'B': 22, 'C': 12, 'D': 14, 'E': 95, 'F': 15, 'G': 45
    })
    max_col = 7

    focus_area = cfg.get('focus_area')
    reviews = cfg.get('third_party_reviews', [])
    grouped = group_reviews_by_platform(reviews)
    total = sum(len(v) for v in grouped.values())

    # Row 1 — Title
    if focus_area:
        title_text = f"{focus_area} — Third-Party Reviews ({total} Reviews)"
    else:
        title_text = f"Third-Party Reviews — {total} Reviews"
    ws.cell(row=1, column=1, value=title_text)
    style_title(ws, 1, max_col)

    row = 2
    for platform, platform_reviews in grouped.items():
        if not platform_reviews:
            continue

        # Platform section header
        ws.cell(
            row=row, column=1,
            value=f"{platform} ({len(platform_reviews)} reviews)"
        )
        style_section(ws, row, max_col)
        row += 1

        # Column headers
        headers = ["#", "Reviewer", "Rating", "Sentiment", "Review Text", "Date", "Source URL"]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=i + 1, value=h)
        style_headers(ws, row, len(headers))
        row += 1

        # Review rows — # resets per platform
        for idx, r in enumerate(platform_reviews, start=1):
            sentiment_val = r.get('sentiment', '')
            font_overrides = [
                None,
                None,
                BOLD_FONT,                          # Rating (bold)
                font_for_sentiment(sentiment_val),  # Sentiment (color)
                None,
                None,
                None,
            ]
            write_data_row(
                ws, row,
                [
                    idx,
                    r.get('reviewer', ''),
                    r.get('rating', ''),
                    sentiment_val,
                    r.get('text', ''),
                    r.get('date', ''),
                    r.get('source_url', ''),
                ],
                font_overrides=font_overrides,
            )
            row += 1


# ---------------------------------------------------------------------------
# Sheet 3: Employee Reviews
# ---------------------------------------------------------------------------

def build_employee_reviews(ws, cfg):
    # Note: C=14, D=12 — swapped widths vs Sheet 2 because Sentiment/Rating swap
    set_column_widths(ws, {
        'A': 5, 'B': 22, 'C': 14, 'D': 12, 'E': 95, 'F': 15, 'G': 45
    })
    max_col = 7

    focus_area = cfg.get('focus_area')
    reviews = cfg.get('employee_reviews', [])
    grouped = group_reviews_by_platform(reviews) if reviews else {}
    total = sum(len(v) for v in grouped.values())

    # Row 1 — Title
    if focus_area:
        title_text = f"Employee Reviews — {focus_area} ({total} Reviews)"
    else:
        title_text = f"Employee Reviews — {total} Reviews"
    ws.cell(row=1, column=1, value=title_text)
    style_title(ws, 1, max_col)

    # If no employee data, leave a notice
    if not grouped:
        ws.cell(row=3, column=1, value="No employee reviews collected for this report.")
        ws.cell(row=3, column=1).font = NORMAL_FONT
        ws.cell(row=3, column=1).alignment = Alignment(vertical='top')
        return

    row = 2
    for platform, platform_reviews in grouped.items():
        if not platform_reviews:
            continue

        # Platform section header
        ws.cell(
            row=row, column=1,
            value=f"{platform} ({len(platform_reviews)} reviews)"
        )
        style_section(ws, row, max_col)
        row += 1

        # Column headers — Sentiment and Rating swapped vs Sheet 2
        headers = ["#", "Reviewer", "Sentiment", "Rating", "Review Text", "Date", "Source URL"]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=i + 1, value=h)
        style_headers(ws, row, len(headers))
        row += 1

        for idx, r in enumerate(platform_reviews, start=1):
            sentiment_val = r.get('sentiment', '')
            font_overrides = [
                None,
                None,
                font_for_sentiment(sentiment_val),  # Sentiment (color) — col C
                BOLD_FONT,                          # Rating (bold) — col D
                None,
                None,
                None,
            ]
            write_data_row(
                ws, row,
                [
                    idx,
                    r.get('reviewer', ''),
                    sentiment_val,
                    r.get('rating', ''),
                    r.get('text', ''),
                    r.get('date', ''),
                    r.get('source_url', ''),
                ],
                font_overrides=font_overrides,
            )
            row += 1


# ---------------------------------------------------------------------------
# Sheet 4: Competitive Benchmarking
# ---------------------------------------------------------------------------

def build_competitive_benchmarking(ws, cfg):
    set_column_widths(ws, {
        'A': 38, 'B': 24, 'C': 24, 'D': 24, 'E': 24, 'F': 58, 'G': 5
    })
    max_col = 7

    company = cfg['company_name']
    prepared_by = cfg['prepared_by']
    report_date = cfg.get('report_date', '')

    # Row 1 — Title
    ws.cell(row=1, column=1, value=f"{company} — Competitive Benchmarking")
    style_title(ws, 1, max_col)

    # Row 2 — Subtitle
    parts = ["Based on competitor review analysis across independent platforms"]
    if report_date:
        parts.append(report_date)
    parts.append(prepared_by)
    ws.cell(row=2, column=1, value=" — ".join([parts[0], " | ".join(parts[1:])]) if len(parts) > 1 else parts[0])
    style_section(ws, 2, max_col)

    row = 3
    tables = cfg.get('competitive_benchmarking', [])
    for table in tables:
        # Section header
        ws.cell(row=row, column=1, value=table.get('title', ''))
        style_section(ws, row, max_col)
        row += 1

        # Column headers
        headers = table.get('headers', [])
        for i, h in enumerate(headers):
            ws.cell(row=row, column=i + 1, value=h)
        style_headers(ws, row, len(headers))
        row += 1

        # Data rows
        for data_row in table.get('rows', []):
            write_data_row(ws, row, data_row)
            row += 1

        row += 1  # gap between tables


# ---------------------------------------------------------------------------
# Sheet 5: Competitor Deep Dives
# ---------------------------------------------------------------------------

def build_competitor_deep_dives(ws, cfg):
    set_column_widths(ws, {
        'A': 8, 'B': 22, 'C': 16, 'D': 12, 'E': 32, 'F': 72, 'G': 14
    })
    max_col = 7

    prepared_by = cfg['prepared_by']
    deep_dives = cfg.get('competitor_deep_dives', [])
    competitor_names = [d.get('competitor', '') for d in deep_dives if d.get('competitor')]

    # Row 1 — Title
    if competitor_names:
        comp_list = ", ".join(competitor_names)
        title_text = f"Competitor Deep Dives — {comp_list}"
    else:
        title_text = "Competitor Deep Dives"
    ws.cell(row=1, column=1, value=title_text)
    style_title(ws, 1, max_col)

    # Row 2 — Subtitle
    ws.cell(
        row=2, column=1,
        value=f"Selected review quotes and complaint themes from independent platforms | {prepared_by}"
    )
    style_section(ws, 2, max_col)

    row = 3
    for dive in deep_dives:
        # Competitor section header
        competitor = dive.get('competitor', '')
        summary_line = dive.get('summary_line', '')
        if summary_line:
            header_text = f"{competitor} — {summary_line}"
        else:
            header_text = competitor
        ws.cell(row=row, column=1, value=header_text)
        style_section(ws, row, max_col)
        row += 1

        # Column headers
        headers = ["#", "Reviewer", "Platform", "Rating", "Theme", "Quote Excerpt", "Date"]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=i + 1, value=h)
        style_headers(ws, row, len(headers))
        row += 1

        # Quote rows
        for idx, q in enumerate(dive.get('quotes', []), start=1):
            font_overrides = [None, None, None, BOLD_FONT, None, None, None]
            write_data_row(
                ws, row,
                [
                    idx,
                    q.get('reviewer', ''),
                    q.get('platform', ''),
                    q.get('rating', ''),
                    q.get('theme', ''),
                    q.get('excerpt', ''),
                    q.get('date', ''),
                ],
                font_overrides=font_overrides,
            )
            row += 1


# ---------------------------------------------------------------------------
# Workbook orchestration
# ---------------------------------------------------------------------------

def build_workbook(cfg, output_path):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = 'Review Summary'
    build_review_summary(ws1, cfg)

    ws2 = wb.create_sheet('Third-Party Reviews')
    build_third_party_reviews(ws2, cfg)

    ws3 = wb.create_sheet('Employee Reviews')
    build_employee_reviews(ws3, cfg)

    ws4 = wb.create_sheet('Competitive Benchmarking')
    build_competitive_benchmarking(ws4, cfg)

    ws5 = wb.create_sheet('Competitor Deep Dives')
    build_competitor_deep_dives(ws5, cfg)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def validate_config(cfg):
    """Light-touch validation. Raise ValueError on missing required fields."""
    required = ['company_name', 'prepared_by']
    missing = [k for k in required if not cfg.get(k)]
    if missing:
        raise ValueError(f"Config missing required fields: {missing}")

    # Soft warnings — print but don't raise
    soft_required = ['platforms', 'themes', 'sentiment_timeline', 'third_party_reviews']
    soft_missing = [k for k in soft_required if not cfg.get(k)]
    if soft_missing:
        print(f"WARNING: config has empty sections (will render but may look thin): {soft_missing}")


def main():
    parser = argparse.ArgumentParser(description="Build a Consumer Intelligence XLSX workbook.")
    parser.add_argument('--config', required=True, help='Path to config JSON.')
    parser.add_argument('--output', required=True, help='Path to output .xlsx file.')
    args = parser.parse_args()

    with open(args.config, 'r', encoding='utf-8') as f:
        cfg = json.load(f)

    validate_config(cfg)
    out = build_workbook(cfg, args.output)
    print(f"Wrote workbook: {out}")


if __name__ == '__main__':
    main()
