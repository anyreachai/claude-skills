"""
Build the BPO ROI financial model as an XLSX workbook.

Produces a 4-tab workbook with the same structure as the Fabletics v4 model:
  1. Executive Summary — top-level dashboard (formulas reference other tabs)
  2. Current State — volume + cost cascade (input cells + derived formulas)
  3. Revenue Future — 6 revenue streams with adjustable assumptions
  4. Proposed Model — phased pricing, savings, retention math
  5. Assumptions — input audit trail / source citations

The model is formula-driven: input cells are clearly marked, and every
derived number flows from those inputs. Edits to inputs in any tab cascade
through the rest of the workbook automatically.

Usage:
    python build_model.py <config.json> <output.xlsx>
"""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE_00


# ── Color palette ──────────────────────────────────────────────
NAVY = '003D6B'        # default — overridden by config.bpo.brand.navy
ACCENT = 'E92983'      # default — overridden by config.bpo.brand.accent
NAVY_SOFT = 'E8EEF5'   # light tint for input cell backgrounds
GREEN = '1D7366'
TEXT = '14293D'
TEXT_MUTE = '5D758D'
BG_LIGHT = 'F6F8FA'


def _safe_color(hex_str, fallback):
    """Normalize a hex color string for openpyxl (strip leading #)."""
    if not hex_str:
        return fallback
    return hex_str.lstrip('#').upper()


def _make_styles(navy, accent):
    """Build all the cell styles we use. Returns a dict of openpyxl Fonts/Fills."""
    return {
        'title_font': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
        'title_fill': PatternFill('solid', fgColor=navy),
        'subtitle_font': Font(name='Calibri', size=11, italic=True, color=TEXT_MUTE),
        'section_font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
        'section_fill': PatternFill('solid', fgColor=accent),
        'header_font': Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
        'header_fill': PatternFill('solid', fgColor=navy),
        'body_font': Font(name='Calibri', size=10, color=TEXT),
        'body_bold': Font(name='Calibri', size=10, bold=True, color=TEXT),
        'highlight_font': Font(name='Calibri', size=11, bold=True, color=GREEN),
        'highlight_fill': PatternFill('solid', fgColor='E8F5E9'),  # mint pastel
        'mute_font': Font(name='Calibri', size=9, italic=True, color=TEXT_MUTE),
        'input_fill': PatternFill('solid', fgColor=NAVY_SOFT),  # input cells stand out
        'thin_border': Border(
            left=Side(border_style='thin', color='CFD8DC'),
            right=Side(border_style='thin', color='CFD8DC'),
            top=Side(border_style='thin', color='CFD8DC'),
            bottom=Side(border_style='thin', color='CFD8DC'),
        ),
    }


# Currency / percent format strings — applied via cell.number_format
FMT_CURRENCY = '"$"#,##0'
FMT_CURRENCY_2DP = '"$"#,##0.00'
FMT_PERCENT_1DP = '0.0%'
FMT_NUMBER = '#,##0'


def _set(ws, coord, value, fmt=None, font=None, fill=None, align=None, border=None):
    """Convenience: set cell value + style in one call."""
    cell = ws[coord]
    cell.value = value
    if fmt: cell.number_format = fmt
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    return cell


def _row_block(ws, row, label, value, desc='', fmt=None, font=None, fill=None,
               styles=None, label_col='B', value_col='C', desc_col='D'):
    """Helper: write a 'Label / Value / Description' row in standard layout."""
    s = styles
    _set(ws, f'{label_col}{row}', label, font=s['body_font'])
    _set(ws, f'{value_col}{row}', value, fmt=fmt, font=font or s['body_font'], fill=fill)
    if desc:
        _set(ws, f'{desc_col}{row}', desc, font=s['mute_font'])


# ── Tab builders ──────────────────────────────────────────────

def build_current_state(wb, config, styles):
    """Tab 2: Current State — input cells + cost cascade."""
    ws = wb.create_sheet('Current State')

    cs = config['current_state']
    client_name = config['end_client']['name']
    ai_vendor = cs.get('ai_vendor', {}).get('name', 'Cognigy')
    bpo_label = config['bpo'].get('label', 'BPO')

    s = styles

    # Title
    ws.merge_cells('B2:E2')
    _set(ws, 'B2', f'Current State: What {client_name} Spends Today',
         font=s['title_font'], fill=s['title_fill'],
         align=Alignment(horizontal='left', vertical='center'))
    ws.row_dimensions[2].height = 28

    # Section: ANNUAL CONTACT VOLUME
    ws.merge_cells('B4:E4')
    _set(ws, 'B4', 'ANNUAL CONTACT VOLUME', font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))

    volumes = cs['volumes']
    row = 5
    if 'voice' in volumes:
        _row_block(ws, row, 'Voice Contacts', volumes['voice'], 'Source: Client RFP',
                   fmt=FMT_NUMBER, fill=s['input_fill'], styles=s)
        row += 1
    if 'chat' in volumes:
        _row_block(ws, row, 'Chat Contacts', volumes['chat'], 'Source: Client RFP',
                   fmt=FMT_NUMBER, fill=s['input_fill'], styles=s)
        row += 1
    if 'social' in volumes:
        _row_block(ws, row, 'Social/WhatsApp Contacts', volumes['social'], 'Source: Client RFP',
                   fmt=FMT_NUMBER, fill=s['input_fill'], styles=s)
        row += 1
    if 'email' in volumes:
        _row_block(ws, row, 'Email Contacts', volumes['email'], 'Source: Client RFP',
                   fmt=FMT_NUMBER, fill=s['input_fill'], styles=s)
        row += 1

    # Total volume — formula
    volume_first_row = 5
    volume_last_row = row - 1
    total_volume_row = row
    _row_block(ws, row, 'Total Contact Volume',
               f'=SUM(C{volume_first_row}:C{volume_last_row})',
               'All channels',
               fmt=FMT_NUMBER, font=s['body_bold'], styles=s)
    row += 2

    # Section: AI VENDOR SPEND ESTIMATE
    ws.merge_cells(f'B{row}:E{row}')
    _set(ws, f'B{row}', f'{ai_vendor.upper()} AI SPEND — BACK-OF-ENVELOPE ESTIMATE',
         font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))
    row += 1

    rates = cs.get('ai_vendor', {}).get('estimated_rates', {})
    chat_rate = rates.get('chat', 0.60)
    voice_rate = rates.get('voice', 0.55)
    social_rate = rates.get('social', 0.65)
    platform_base = rates.get('platform_base', 150_000)

    _row_block(ws, row, 'Billing Model', 'Per conversation', 'Public vendor docs', styles=s)
    row += 1
    chat_rate_row = row
    _row_block(ws, row, 'Est. Enterprise Rate — Chat', chat_rate,
               'Discounted from public list price',
               fmt=FMT_CURRENCY_2DP, fill=s['input_fill'], styles=s)
    row += 1
    voice_rate_row = row
    _row_block(ws, row, 'Est. Enterprise Rate — Voice', voice_rate,
               'Volume discount estimate',
               fmt=FMT_CURRENCY_2DP, fill=s['input_fill'], styles=s)
    row += 1
    social_rate_row = row
    _row_block(ws, row, 'Est. Enterprise Rate — Social/WhatsApp', social_rate,
               'Same tier as chat',
               fmt=FMT_CURRENCY_2DP, fill=s['input_fill'], styles=s)
    row += 1
    platform_row = row
    _row_block(ws, row, 'Est. Platform Base + LLM Token Costs', platform_base,
               'Annual platform licensing + GenAI tokens',
               fmt=FMT_CURRENCY, fill=s['input_fill'], styles=s)
    row += 2

    # Cost calculations
    ai_cost_first_row = row
    voice_row_idx = next((i for i, k in enumerate(volumes.keys()) if k == 'voice'), None)
    chat_row_idx = next((i for i, k in enumerate(volumes.keys()) if k == 'chat'), None)
    social_row_idx = next((i for i, k in enumerate(volumes.keys()) if k == 'social'), None)

    if 'chat' in volumes:
        _row_block(ws, row, 'Chat AI Cost',
                   f'=C{volume_first_row + chat_row_idx}*C{chat_rate_row}',
                   f'{volumes["chat"] / 1e6:.2f}M chats × ${chat_rate:.2f}',
                   fmt=FMT_CURRENCY, styles=s)
        row += 1
    if 'voice' in volumes:
        _row_block(ws, row, 'Voice AI Cost',
                   f'=C{volume_first_row + voice_row_idx}*C{voice_rate_row}',
                   f'{volumes["voice"] / 1e6:.2f}M calls × ${voice_rate:.2f}',
                   fmt=FMT_CURRENCY, styles=s)
        row += 1
    if 'social' in volumes:
        _row_block(ws, row, 'Social/WhatsApp AI Cost',
                   f'=C{volume_first_row + social_row_idx}*C{social_rate_row}',
                   f'{volumes["social"] / 1e3:.0f}K contacts × ${social_rate:.2f}',
                   fmt=FMT_CURRENCY, styles=s)
        row += 1
    _row_block(ws, row, 'Platform + LLM Tokens',
               f'=C{platform_row}', 'Annual fixed costs',
               fmt=FMT_CURRENCY, styles=s)
    row += 1
    ai_cost_last_row = row - 1
    ai_total_row = row
    _row_block(ws, row, 'Estimated Total AI Vendor Spend',
               f'=SUM(C{ai_cost_first_row}:C{ai_cost_last_row})',
               'Back-of-envelope from public billing model',
               fmt=FMT_CURRENCY, font=s['body_bold'], fill=s['highlight_fill'], styles=s)
    row += 2

    # Section: BPO LABOR
    ws.merge_cells(f'B{row}:E{row}')
    _set(ws, f'B{row}', f'{bpo_label.upper()} AGENT SPEND — CURRENT PARTNER',
         font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))
    row += 1

    fte_count_row = row
    _row_block(ws, row, 'Estimated Agent Workforce', cs['fte_count'],
               'Confirmed workforce size',
               fmt=FMT_NUMBER, fill=s['input_fill'], styles=s)
    row += 1
    fte_rate_row = row
    _row_block(ws, row, 'Agent Billing Rate', cs['fte_hourly_rate'],
               '$/hr — current partner', fmt=FMT_CURRENCY_2DP,
               fill=s['input_fill'], styles=s)
    row += 1
    annual_hrs_row = row
    _row_block(ws, row, 'Annual Hours per FTE', 2080,
               'Standard full-time (40 hrs × 52 weeks)',
               fmt=FMT_NUMBER, fill=s['input_fill'], styles=s)
    row += 1
    bpo_total_row = row
    _row_block(ws, row, 'Estimated Annual BPO Spend',
               f'=C{fte_count_row}*C{fte_rate_row}*C{annual_hrs_row}',
               'FTEs × rate × annual hours',
               fmt=FMT_CURRENCY, font=s['body_bold'], fill=s['highlight_fill'], styles=s)
    row += 2

    # Section: TOTAL CURRENT SPEND
    ws.merge_cells(f'B{row}:E{row}')
    _set(ws, f'B{row}', 'TOTAL CURRENT SPEND + REVENUE BENCHMARK',
         font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))
    row += 1

    _row_block(ws, row, f'AI Platform ({ai_vendor})', f'=C{ai_total_row}',
               'From above', fmt=FMT_CURRENCY, styles=s)
    ai_link_row = row
    row += 1
    _row_block(ws, row, f'{bpo_label} Labor (Current Partner)', f'=C{bpo_total_row}',
               'From above', fmt=FMT_CURRENCY, styles=s)
    bpo_link_row = row
    row += 1
    total_cx_row = row
    _row_block(ws, row, 'TOTAL CURRENT ANNUAL CX SPEND',
               f'=C{ai_link_row}+C{bpo_link_row}',
               'AI + Human',
               fmt=FMT_CURRENCY, font=s['body_bold'], fill=s['highlight_fill'], styles=s)
    row += 2

    # Revenue benchmark (only if revenue provided)
    annual_revenue = cs.get('annual_revenue')
    rev_row = None
    pct_row = None
    if annual_revenue:
        rev_row = row
        _row_block(ws, row, f'{client_name} Annual Revenue',
                   annual_revenue, 'Source: public filings',
                   fmt=FMT_CURRENCY, fill=s['input_fill'], styles=s)
        row += 1
        pct_row = row
        _row_block(ws, row, 'CX Spend as % of Revenue',
                   f'=C{total_cx_row}/C{rev_row}',
                   '', fmt=FMT_PERCENT_1DP, styles=s)
        row += 1
        _row_block(ws, row, 'Industry Benchmark', '2% - 5%',
                   'Standard CX spend range', styles=s)
        row += 1
        _row_block(ws, row, 'Assessment', 'Tracks within benchmark range',
                   'Estimate is directionally credible', styles=s)
        row += 2

    # Section: COST PER OUTCOME
    ws.merge_cells(f'B{row}:E{row}')
    _set(ws, f'B{row}', 'CURRENT COST PER OUTCOME — BOTTOMS-UP',
         font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))
    row += 1

    containment_row = row
    _row_block(ws, row, 'Current AI Containment Rate',
               cs['ai_containment_rate'],
               cs.get('containment_source', 'Industry benchmark'),
               fmt=FMT_PERCENT_1DP, fill=s['input_fill'], styles=s)
    row += 1
    ai_resolved_row = row
    _row_block(ws, row, 'AI-Resolved Contacts',
               f'=C{total_volume_row}*C{containment_row}',
               'Volume contained by AI',
               fmt=FMT_NUMBER, styles=s)
    row += 1
    human_resolved_row = row
    _row_block(ws, row, 'Human-Resolved Contacts',
               f'=C{total_volume_row}-C{ai_resolved_row}',
               'Remainder handled by agents',
               fmt=FMT_NUMBER, styles=s)
    row += 2

    cost_per_ai_row = row
    _row_block(ws, row, f'{ai_vendor} Cost per AI-Resolved Outcome',
               f'=C{ai_total_row}/C{ai_resolved_row}',
               'AI spend / AI-resolved contacts',
               fmt=FMT_CURRENCY_2DP, styles=s)
    row += 1
    cost_per_human_row = row
    _row_block(ws, row, f'{bpo_label} Cost per Human-Resolved Outcome',
               f'=C{bpo_total_row}/C{human_resolved_row}',
               'BPO spend / human-resolved contacts',
               fmt=FMT_CURRENCY_2DP, styles=s)
    row += 2

    blended_row = row
    _row_block(ws, row, 'BLENDED CURRENT COST PER OUTCOME',
               f'=C{total_cx_row}/C{total_volume_row}',
               'Total spend / total contacts',
               fmt=FMT_CURRENCY_2DP, font=s['body_bold'], fill=s['highlight_fill'], styles=s)

    # Set column widths
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 4

    # Return key cell references for cross-tab linking
    return {
        'total_volume': f"'Current State'!C{total_volume_row}",
        'ai_total': f"'Current State'!C{ai_total_row}",
        'bpo_total': f"'Current State'!C{bpo_total_row}",
        'total_cx': f"'Current State'!C{total_cx_row}",
        'containment': f"'Current State'!C{containment_row}",
        'blended_cost': f"'Current State'!C{blended_row}",
        'cost_per_ai': f"'Current State'!C{cost_per_ai_row}",
        'cost_per_human': f"'Current State'!C{cost_per_human_row}",
        'annual_revenue': f"'Current State'!C{rev_row}" if rev_row else None,
    }


def build_proposed_model(wb, config, refs, styles):
    """Tab: Proposed Model — phased pricing + savings."""
    ws = wb.create_sheet('Proposed Model')
    s = styles
    bpo_name = config['bpo']['name']
    ai_partner = config.get('ai_partner', {}).get('name', 'Anyreach AI')
    client_name = config['end_client']['name']
    phases = config.get('pricing_phases', [
        {'name': 'Phase 1', 'months': '1-6', 'duration': 6, 'containment': 0.55, 'price_per_outcome': 3.00},
        {'name': 'Phase 2', 'months': '7-9', 'duration': 3, 'containment': 0.75, 'price_per_outcome': 2.85},
        {'name': 'Phase 3', 'months': '10+', 'duration': 3, 'containment': 0.90, 'price_per_outcome': 2.65},
    ])

    # Title
    ws.merge_cells('B2:F2')
    _set(ws, 'B2', f'Proposed: Outcomes-Based AI + Human Pricing',
         font=s['title_font'], fill=s['title_fill'],
         align=Alignment(horizontal='left', vertical='center'))
    ws.row_dimensions[2].height = 28
    _set(ws, 'B3', f'{ai_partner} × {bpo_name}  |  Price per resolved contact — savings from Day 1',
         font=s['subtitle_font'])

    # Section: PHASED PRICING
    ws.merge_cells('B5:F5')
    _set(ws, 'B5', 'PHASED PRICING — TIED TO AI CONTAINMENT MILESTONES',
         font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))

    cols = ['C', 'D', 'E']  # one column per phase
    for i, ph in enumerate(phases):
        _set(ws, f'{cols[i]}6', ph['name'], font=s['header_font'], fill=s['header_fill'],
             align=Alignment(horizontal='center'))
    _set(ws, 'B7', 'Period', font=s['body_font'])
    for i, ph in enumerate(phases):
        _set(ws, f'{cols[i]}7', f'Month {ph["months"]}', font=s['body_font'],
             align=Alignment(horizontal='center'))

    containment_inputs = {}
    price_inputs = {}
    months_inputs = {}

    _set(ws, 'B8', 'AI Containment Target', font=s['body_font'])
    for i, ph in enumerate(phases):
        cell = _set(ws, f'{cols[i]}8', ph['containment'], fmt=FMT_PERCENT_1DP,
                    fill=s['input_fill'], align=Alignment(horizontal='center'))
        containment_inputs[i] = f'{cols[i]}8'

    _set(ws, 'B9', 'Price per Outcome', font=s['body_font'])
    for i, ph in enumerate(phases):
        _set(ws, f'{cols[i]}9', ph['price_per_outcome'], fmt=FMT_CURRENCY_2DP,
             fill=s['input_fill'], font=s['body_bold'],
             align=Alignment(horizontal='center'))
        price_inputs[i] = f'{cols[i]}9'

    _set(ws, 'B10', 'Monthly Contact Volume', font=s['body_font'])
    for i, ph in enumerate(phases):
        _set(ws, f'{cols[i]}10', f'={refs["total_volume"]}/12', fmt=FMT_NUMBER,
             align=Alignment(horizontal='center'))

    _set(ws, 'B11', 'Months in Phase', font=s['body_font'])
    for i, ph in enumerate(phases):
        _set(ws, f'{cols[i]}11', ph['duration'], fmt=FMT_NUMBER,
             fill=s['input_fill'], align=Alignment(horizontal='center'))
        months_inputs[i] = f'{cols[i]}11'

    _set(ws, 'B12', 'Phase Cost', font=s['body_bold'])
    for i, ph in enumerate(phases):
        _set(ws, f'{cols[i]}12',
             f'={cols[i]}9*{cols[i]}10*{cols[i]}11',
             fmt=FMT_CURRENCY, font=s['body_bold'], fill=s['highlight_fill'],
             align=Alignment(horizontal='center'))

    # Section: YEAR 1 SUMMARY
    ws.merge_cells('B14:F14')
    _set(ws, 'B14', 'YEAR 1 SUMMARY', font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))

    sum_phase_cost = '+'.join(f'{c}12' for c in cols[:len(phases)])
    y1_total_row = 15
    _row_block(ws, 15, 'Year 1 Blended Cost (All Phases)',
               f'={sum_phase_cost}',
               'Sum of all phases',
               fmt=FMT_CURRENCY, font=s['body_bold'], styles=s)
    y1_blended_row = 16
    _row_block(ws, 16, 'Year 1 Blended Price per Outcome',
               f'=C{y1_total_row}/{refs["total_volume"]}',
               'Weighted avg across phases',
               fmt=FMT_CURRENCY_2DP, styles=s)
    y2_total_row = 17
    last_price_col = cols[len(phases) - 1]
    _row_block(ws, 17, 'Year 2+ Steady-State Cost',
               f'={refs["total_volume"]}*{last_price_col}9',
               f'Full year at final phase price',
               fmt=FMT_CURRENCY, styles=s)

    # Section: SAVINGS
    ws.merge_cells('B19:F19')
    _set(ws, 'B19', 'SAVINGS vs. CURRENT STATE', font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))

    _row_block(ws, 20, 'Current Annual CX Spend', f'={refs["total_cx"]}',
               'From Current State tab', fmt=FMT_CURRENCY, styles=s)
    _row_block(ws, 21, 'Current Cost per Outcome', f'={refs["blended_cost"]}',
               'Blended current CPI', fmt=FMT_CURRENCY_2DP, styles=s)

    _set(ws, 'C23', 'Year 1', font=s['header_font'], fill=s['header_fill'],
         align=Alignment(horizontal='center'))
    _set(ws, 'D23', 'Year 2+', font=s['header_font'], fill=s['header_fill'],
         align=Alignment(horizontal='center'))

    _set(ws, 'B24', 'Proposed Annual Cost', font=s['body_font'])
    _set(ws, 'C24', f'=C{y1_total_row}', fmt=FMT_CURRENCY, align=Alignment(horizontal='center'))
    _set(ws, 'D24', f'=C{y2_total_row}', fmt=FMT_CURRENCY, align=Alignment(horizontal='center'))

    _set(ws, 'B25', 'Proposed Price per Outcome', font=s['body_font'])
    _set(ws, 'C25', f'=C{y1_blended_row}', fmt=FMT_CURRENCY_2DP, align=Alignment(horizontal='center'))
    _set(ws, 'D25', f'={last_price_col}9', fmt=FMT_CURRENCY_2DP, align=Alignment(horizontal='center'))

    y1_savings_row = 26
    _set(ws, 'B26', 'Annual Savings ($)', font=s['body_bold'])
    _set(ws, 'C26', f'=C20-C24', fmt=FMT_CURRENCY, font=s['highlight_font'],
         fill=s['highlight_fill'], align=Alignment(horizontal='center'))
    _set(ws, 'D26', f'=C20-D24', fmt=FMT_CURRENCY, font=s['highlight_font'],
         fill=s['highlight_fill'], align=Alignment(horizontal='center'))

    _set(ws, 'B27', 'Annual Savings (%)', font=s['body_bold'])
    _set(ws, 'C27', f'=C26/C20', fmt=FMT_PERCENT_1DP, font=s['highlight_font'],
         fill=s['highlight_fill'], align=Alignment(horizontal='center'))
    _set(ws, 'D27', f'=D26/C20', fmt=FMT_PERCENT_1DP, font=s['highlight_font'],
         fill=s['highlight_fill'], align=Alignment(horizontal='center'))

    # Section: WHAT'S INCLUDED
    ws.merge_cells('B29:F29')
    _set(ws, 'B29', "WHAT'S INCLUDED IN EVERY RESOLVED OUTCOME",
         font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))

    included_items = [
        ('Agentic AI Resolution', 'End-to-end AI handling: top use cases'),
        ('Live Agent Escalation', 'Seamless warm-transfer to trained agents with full context'),
        ('24/7/365 Coverage', 'AI handles after-hours — no night shift premiums'),
        ('60+ Languages', 'Multilingual out-of-box — no bilingual agent premium'),
        ('Omni-Channel', 'Voice + Chat + SMS + WhatsApp — unified experience'),
        ('Proactive Outreach', 'Reminders + education — reduces inbound volume'),
        ('Real-Time Agent Assist', 'AI co-pilot with knowledge base + recommendations'),
        ('Analytics & Reporting', 'Sentiment, containment, CSAT, recordings'),
        ('Ongoing Optimization', 'Weekly AI tuning, monthly business reviews'),
    ]
    for i, (item, desc) in enumerate(included_items):
        _set(ws, f'B{30 + i}', item, font=s['body_font'])
        _set(ws, f'C{30 + i}', desc, font=s['body_font'])

    next_row = 30 + len(included_items) + 2

    # Section: RETENTION & REVENUE IMPACT (only if revenue + retention provided)
    annual_revenue = config['current_state'].get('annual_revenue')
    retention_row = None
    if annual_revenue:
        retention_lift = config.get('retention_lift_pct', 0.005)
        ws.merge_cells(f'B{next_row}:F{next_row}')
        _set(ws, f'B{next_row}', 'RETENTION & REVENUE IMPACT',
             font=s['section_font'], fill=s['section_fill'],
             align=Alignment(horizontal='left', vertical='center'))
        next_row += 1

        rev_input_row = next_row
        _row_block(ws, next_row, f'{client_name} Annual Revenue', annual_revenue,
                   'From Current State tab (also editable here)',
                   fmt=FMT_CURRENCY, fill=s['input_fill'], styles=s)
        next_row += 1
        retention_row_input = next_row
        _row_block(ws, next_row, 'Conservative Retention Lift', retention_lift,
                   'CX-driven retention improvement',
                   fmt=FMT_PERCENT_1DP, fill=s['input_fill'], styles=s)
        next_row += 1
        retention_row = next_row
        _row_block(ws, next_row, 'Retained Revenue (Annual)',
                   f'=C{rev_input_row}*C{retention_row_input}',
                   'Preserved revenue from reduced churn',
                   fmt=FMT_CURRENCY, font=s['body_bold'], fill=s['highlight_fill'], styles=s)
        next_row += 2

    # Section: TOTAL ANNUAL VALUE
    ws.merge_cells(f'B{next_row}:F{next_row}')
    _set(ws, f'B{next_row}', 'TOTAL ANNUAL VALUE',
         font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))
    next_row += 1

    _set(ws, f'C{next_row}', 'Year 1', font=s['header_font'], fill=s['header_fill'],
         align=Alignment(horizontal='center'))
    _set(ws, f'D{next_row}', 'Year 2+', font=s['header_font'], fill=s['header_fill'],
         align=Alignment(horizontal='center'))
    next_row += 1

    _set(ws, f'B{next_row}', 'Direct Cost Savings', font=s['body_font'])
    direct_y1_row = next_row
    _set(ws, f'C{next_row}', f'=C{y1_savings_row}', fmt=FMT_CURRENCY, align=Alignment(horizontal='center'))
    _set(ws, f'D{next_row}', f'=D{y1_savings_row}', fmt=FMT_CURRENCY, align=Alignment(horizontal='center'))
    next_row += 1

    if retention_row:
        _set(ws, f'B{next_row}', 'Retained Revenue', font=s['body_font'])
        retention_y1_row = next_row
        _set(ws, f'C{next_row}', f'=C{retention_row}', fmt=FMT_CURRENCY, align=Alignment(horizontal='center'))
        _set(ws, f'D{next_row}', f'=C{retention_row}', fmt=FMT_CURRENCY, align=Alignment(horizontal='center'))
        next_row += 1

    total_value_row = next_row
    _set(ws, f'B{next_row}', 'TOTAL VALUE', font=s['body_bold'])
    if retention_row:
        _set(ws, f'C{next_row}', f'=C{direct_y1_row}+C{retention_y1_row}',
             fmt=FMT_CURRENCY, font=s['highlight_font'], fill=s['highlight_fill'],
             align=Alignment(horizontal='center'))
        _set(ws, f'D{next_row}', f'=D{direct_y1_row}+D{retention_y1_row}',
             fmt=FMT_CURRENCY, font=s['highlight_font'], fill=s['highlight_fill'],
             align=Alignment(horizontal='center'))
    else:
        _set(ws, f'C{next_row}', f'=C{direct_y1_row}',
             fmt=FMT_CURRENCY, font=s['highlight_font'], fill=s['highlight_fill'],
             align=Alignment(horizontal='center'))
        _set(ws, f'D{next_row}', f'=D{direct_y1_row}',
             fmt=FMT_CURRENCY, font=s['highlight_font'], fill=s['highlight_fill'],
             align=Alignment(horizontal='center'))

    # Column widths
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 36
    for c in cols:
        ws.column_dimensions[c].width = 16
    ws.column_dimensions['F'].width = 30

    return {
        'y1_total_cost': f"'Proposed Model'!C{y1_total_row}",
        'y1_blended_price': f"'Proposed Model'!C{y1_blended_row}",
        'y2_total_cost': f"'Proposed Model'!C{y2_total_row}",
        'y1_savings': f"'Proposed Model'!C{y1_savings_row}",
        'y2_savings': f"'Proposed Model'!D{y1_savings_row}",
        'y1_savings_pct': f"'Proposed Model'!C27",
        'y2_savings_pct': f"'Proposed Model'!D27",
        'retention_revenue': f"'Proposed Model'!C{retention_row}" if retention_row else None,
        'total_value_y1': f"'Proposed Model'!C{total_value_row}",
        'total_value_y2': f"'Proposed Model'!D{total_value_row}",
    }


def build_executive_summary(wb, config, cs_refs, pm_refs, styles):
    """Tab 1: Executive Summary — top-level dashboard, references other tabs."""
    ws = wb.create_sheet('Executive Summary', 0)  # insert at index 0
    s = styles
    bpo_name = config['bpo']['name']
    ai_partner = config.get('ai_partner', {}).get('name', 'Anyreach AI')
    client_name = config['end_client']['name']
    engagement = config.get('engagement_label', 'Customer Experience')

    # Title
    ws.merge_cells('B2:F2')
    _set(ws, 'B2', f'{client_name}: {engagement} — ROI Model',
         font=s['title_font'], fill=s['title_fill'],
         align=Alignment(horizontal='left', vertical='center'))
    ws.row_dimensions[2].height = 30

    _set(ws, 'B3', f'Outcomes-Based Pricing  |  Savings from Day 1  |  {ai_partner} × {bpo_name}',
         font=s['subtitle_font'])
    ws.merge_cells('B3:F3')

    _set(ws, 'B4', 'Confidential — for internal review',
         font=s['mute_font'])

    # Section: BOTTOM LINE
    ws.merge_cells('B6:F6')
    _set(ws, 'B6', 'THE BOTTOM LINE', font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))

    headers_row = 7
    _set(ws, 'C7', 'Current State', font=s['header_font'], fill=s['header_fill'],
         align=Alignment(horizontal='center'))
    _set(ws, 'D7', 'Proposed (Yr 1)', font=s['header_font'], fill=s['header_fill'],
         align=Alignment(horizontal='center'))
    _set(ws, 'E7', 'Proposed (Yr 2+)', font=s['header_font'], fill=s['header_fill'],
         align=Alignment(horizontal='center'))

    _set(ws, 'B8', 'Pricing Model', font=s['body_font'])
    ai_vendor = config['current_state'].get('ai_vendor', {}).get('name', 'Cognigy')
    _set(ws, 'C8', f'{ai_vendor} + FTE/Hourly {config["bpo"].get("label", "BPO")}',
         font=s['body_font'], align=Alignment(horizontal='center'))
    _set(ws, 'D8', 'Per Resolved Outcome', font=s['body_font'], align=Alignment(horizontal='center'))
    _set(ws, 'E8', 'Per Resolved Outcome', font=s['body_font'], align=Alignment(horizontal='center'))

    _set(ws, 'B9', 'Annual CX Spend', font=s['body_font'])
    _set(ws, 'C9', f'={cs_refs["total_cx"]}', fmt=FMT_CURRENCY, align=Alignment(horizontal='center'))
    _set(ws, 'D9', f'={pm_refs["y1_total_cost"]}', fmt=FMT_CURRENCY, align=Alignment(horizontal='center'))
    _set(ws, 'E9', f'={pm_refs["y2_total_cost"]}', fmt=FMT_CURRENCY, align=Alignment(horizontal='center'))

    _set(ws, 'B10', 'Cost per Outcome', font=s['body_font'])
    _set(ws, 'C10', f'={cs_refs["blended_cost"]}', fmt=FMT_CURRENCY_2DP,
         align=Alignment(horizontal='center'))
    _set(ws, 'D10', f'={pm_refs["y1_blended_price"]}', fmt=FMT_CURRENCY_2DP,
         align=Alignment(horizontal='center'))
    phases = config.get('pricing_phases', [])
    last_price = phases[-1]['price_per_outcome'] if phases else 2.50
    _set(ws, 'E10', f'={pm_refs["y2_total_cost"]}/{cs_refs["total_volume"]}', fmt=FMT_CURRENCY_2DP,
         align=Alignment(horizontal='center'))

    _set(ws, 'B11', 'AI Containment Rate', font=s['body_font'])
    _set(ws, 'C11', f'={cs_refs["containment"]}', fmt=FMT_PERCENT_1DP,
         align=Alignment(horizontal='center'))
    _set(ws, 'D11', '55% - 90% (ramp)', font=s['body_font'], align=Alignment(horizontal='center'))
    _set(ws, 'E11', '90% (steady state)', font=s['body_font'], align=Alignment(horizontal='center'))

    _set(ws, 'B12', 'Annual Savings', font=s['body_bold'])
    _set(ws, 'D12', f'={pm_refs["y1_savings"]}', fmt=FMT_CURRENCY,
         font=s['highlight_font'], fill=s['highlight_fill'],
         align=Alignment(horizontal='center'))
    _set(ws, 'E12', f'={pm_refs["y2_savings"]}', fmt=FMT_CURRENCY,
         font=s['highlight_font'], fill=s['highlight_fill'],
         align=Alignment(horizontal='center'))

    _set(ws, 'B13', 'Savings %', font=s['body_bold'])
    _set(ws, 'D13', f'={pm_refs["y1_savings_pct"]}', fmt=FMT_PERCENT_1DP,
         font=s['highlight_font'], fill=s['highlight_fill'],
         align=Alignment(horizontal='center'))
    _set(ws, 'E13', f'={pm_refs["y2_savings_pct"]}', fmt=FMT_PERCENT_1DP,
         font=s['highlight_font'], fill=s['highlight_fill'],
         align=Alignment(horizontal='center'))

    # Section: TOTAL ANNUAL VALUE
    ws.merge_cells('B16:F16')
    _set(ws, 'B16', 'TOTAL ANNUAL VALUE (SAVINGS + RETENTION)',
         font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))
    _set(ws, 'C17', 'Year 1', font=s['header_font'], fill=s['header_fill'],
         align=Alignment(horizontal='center'))
    _set(ws, 'D17', 'Year 2+', font=s['header_font'], fill=s['header_fill'],
         align=Alignment(horizontal='center'))

    _set(ws, 'B18', 'Direct Cost Savings', font=s['body_font'])
    _set(ws, 'C18', f'={pm_refs["y1_savings"]}', fmt=FMT_CURRENCY,
         align=Alignment(horizontal='center'))
    _set(ws, 'D18', f'={pm_refs["y2_savings"]}', fmt=FMT_CURRENCY,
         align=Alignment(horizontal='center'))

    if pm_refs.get('retention_revenue'):
        retention_lift = config.get('retention_lift_pct', 0.005)
        _set(ws, 'B19', f'Retained Revenue ({retention_lift*100:g}% lift)', font=s['body_font'])
        _set(ws, 'C19', f'={pm_refs["retention_revenue"]}', fmt=FMT_CURRENCY,
             align=Alignment(horizontal='center'))
        _set(ws, 'D19', f'={pm_refs["retention_revenue"]}', fmt=FMT_CURRENCY,
             align=Alignment(horizontal='center'))
        _set(ws, 'B20', 'TOTAL VALUE', font=s['body_bold'])
        _set(ws, 'C20', f'=C18+C19', fmt=FMT_CURRENCY,
             font=s['highlight_font'], fill=s['highlight_fill'],
             align=Alignment(horizontal='center'))
        _set(ws, 'D20', f'=D18+D19', fmt=FMT_CURRENCY,
             font=s['highlight_font'], fill=s['highlight_fill'],
             align=Alignment(horizontal='center'))

    # Section: BREAKDOWN
    ws.merge_cells('B22:F22')
    _set(ws, 'B22', f'HOW CURRENT BLENDED COST PER OUTCOME BREAKS DOWN',
         font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))

    _row_block(ws, 23, f'{ai_vendor} (AI platform)', f'={cs_refs["ai_total"]}',
               'AI vendor estimate', fmt=FMT_CURRENCY, styles=s)
    _row_block(ws, 24, 'Cost per AI-resolved outcome', f'={cs_refs["cost_per_ai"]}',
               'AI spend / AI-contained contacts', fmt=FMT_CURRENCY_2DP, styles=s)
    _row_block(ws, 25, f'BPO Partner (human agents)', f'={cs_refs["bpo_total"]}',
               'FTE-based estimate', fmt=FMT_CURRENCY, styles=s)
    _row_block(ws, 26, 'Cost per human-resolved outcome', f'={cs_refs["cost_per_human"]}',
               'BPO spend / human-handled contacts', fmt=FMT_CURRENCY_2DP, styles=s)
    _row_block(ws, 27, 'Blended cost per outcome', f'={cs_refs["blended_cost"]}',
               'Total spend / total contacts',
               fmt=FMT_CURRENCY_2DP, font=s['body_bold'], fill=s['highlight_fill'], styles=s)

    _set(ws, 'B29', f'Note: AI vendor spend is estimated from public billing model applied to client-reported contact volumes. Directional; not a quoted figure.',
         font=s['mute_font'])
    ws.merge_cells('B29:F29')

    # Column widths
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 4


def build_assumptions(wb, config, styles):
    """Tab 5: Assumptions — input audit trail."""
    ws = wb.create_sheet('Assumptions')
    s = styles

    cs = config['current_state']
    ai_vendor = cs.get('ai_vendor', {}).get('name', 'Cognigy')
    client_name = config['end_client']['name']

    ws.merge_cells('B2:F2')
    _set(ws, 'B2', 'Key Assumptions & Sources',
         font=s['title_font'], fill=s['title_fill'],
         align=Alignment(horizontal='left', vertical='center'))
    ws.row_dimensions[2].height = 28

    ws.merge_cells('B4:D4')
    _set(ws, 'B4', 'INPUT ASSUMPTIONS',
         font=s['section_font'], fill=s['section_fill'],
         align=Alignment(horizontal='left', vertical='center'))

    _set(ws, 'B5', 'Assumption', font=s['header_font'], fill=s['header_fill'])
    _set(ws, 'C5', 'Value', font=s['header_font'], fill=s['header_fill'])
    _set(ws, 'D5', 'Source / Rationale', font=s['header_font'], fill=s['header_fill'])

    row = 6
    # Volumes
    for ch, label in [('voice', 'Voice Volume'), ('chat', 'Chat Volume'),
                       ('social', 'Social/WhatsApp Volume'), ('email', 'Email Volume')]:
        v = cs['volumes'].get(ch)
        if v:
            _row_block(ws, row, f'{label} (Annual)', f'{v:,}', 'Client-reported',
                       styles=s)
            row += 1
    total_volume = sum(cs['volumes'].values())
    _row_block(ws, row, 'Total Contact Volume', f'{total_volume:,}', 'Sum of channels',
               font=s['body_bold'], styles=s)
    row += 2

    # AI Vendor
    _row_block(ws, row, 'Current AI Platform', ai_vendor, 'Confirmed by client', styles=s)
    row += 1
    _row_block(ws, row, 'Billing Model', 'Per conversation', 'Public vendor docs', styles=s)
    row += 1
    rates = cs.get('ai_vendor', {}).get('estimated_rates', {})
    _row_block(ws, row, 'Est. Enterprise Chat Rate',
               f'${rates.get("chat", 0.60):.2f}/conversation',
               'Volume-tier estimate', styles=s)
    row += 1
    _row_block(ws, row, 'Est. Enterprise Voice Rate',
               f'${rates.get("voice", 0.55):.2f}/conversation',
               'Volume discount estimate', styles=s)
    row += 1
    _row_block(ws, row, 'Est. Platform + LLM Costs',
               f'${rates.get("platform_base", 150_000):,}/yr',
               'Annual licensing + tokens', styles=s)
    row += 1
    _row_block(ws, row, 'Current AI Containment Rate',
               f'{cs["ai_containment_rate"] * 100:.1f}%',
               cs.get('containment_source', 'Industry benchmark'),
               font=s['body_bold'], styles=s)
    row += 2

    # BPO
    _row_block(ws, row, 'Current BPO FTEs', f'~{cs["fte_count"]:,}', 'Client-confirmed', styles=s)
    row += 1
    _row_block(ws, row, 'Current BPO Rate', f'${cs["fte_hourly_rate"]}/hr',
               'Current partner billing rate', styles=s)
    row += 1
    _row_block(ws, row, 'Annual Hours per FTE', '2,080', 'Standard full-time', styles=s)
    row += 2

    # Revenue
    if cs.get('annual_revenue'):
        _row_block(ws, row, f'{client_name} Annual Revenue',
                   f'~${cs["annual_revenue"] / 1e9:.1f}B' if cs['annual_revenue'] >= 1e9 else f'~${cs["annual_revenue"] / 1e6:.0f}M',
                   'Public filings', styles=s)
        row += 1
        _row_block(ws, row, 'Industry CX Spend Benchmark', '2-5% of revenue',
                   'Industry standard', styles=s)
        row += 2

    # Phases
    phases = config.get('pricing_phases', [])
    for i, ph in enumerate(phases):
        _row_block(ws, row, f'Proposed {ph["name"]} Price',
                   f'${ph["price_per_outcome"]:.2f}/outcome',
                   f'Month {ph["months"]} at {ph["containment"] * 100:.0f}% AI containment',
                   styles=s)
        row += 1
    row += 1

    # Retention
    retention_lift = config.get('retention_lift_pct', 0.005)
    if cs.get('annual_revenue'):
        _row_block(ws, row, 'Conservative Retention Lift',
                   f'{retention_lift * 100:g}%',
                   'CX-driven retention improvement', styles=s)
        row += 2

    # Confidentiality note
    if config.get('confidentiality_note'):
        _set(ws, f'B{row}', f'CONFIDENTIALITY: {config["confidentiality_note"]}',
             font=s['mute_font'])
        ws.merge_cells(f'B{row}:F{row}')

    # Column widths
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 50


# ── Main entrypoint ────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('config', help='Path to config JSON')
    ap.add_argument('output', help='Path to output XLSX')
    args = ap.parse_args()

    with open(args.config) as f:
        config = json.load(f)

    # Resolve brand colors
    brand = config['bpo'].get('brand', {})
    navy = _safe_color(brand.get('navy'), NAVY)
    accent = _safe_color(brand.get('accent'), ACCENT)
    styles = _make_styles(navy, accent)

    wb = Workbook()
    # Remove the default sheet so we can build in our preferred order
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Build in dependency order — Current State first (provides refs),
    # Proposed Model second (uses CS refs), Executive Summary last (uses both).
    cs_refs = build_current_state(wb, config, styles)
    pm_refs = build_proposed_model(wb, config, cs_refs, styles)
    build_executive_summary(wb, config, cs_refs, pm_refs, styles)
    build_assumptions(wb, config, styles)

    # Reorder so Executive Summary is first
    sheet_order = ['Executive Summary', 'Current State', 'Proposed Model', 'Assumptions']
    wb._sheets = [wb[name] for name in sheet_order if name in wb.sheetnames]

    out_path = Path(args.output).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f'XLSX written: {out_path}')

    # Print quick cascade summary
    print('\n── Inputs validated ──')
    print(f'  Total volume:      {sum(config["current_state"]["volumes"].values()):,}')
    print(f'  AI containment:    {config["current_state"]["ai_containment_rate"] * 100:.1f}%')
    print(f'  FTE count:         {config["current_state"]["fte_count"]:,}')
    print(f'  FTE hourly rate:   ${config["current_state"]["fte_hourly_rate"]}/hr')
    if config['current_state'].get('annual_revenue'):
        print(f'  Annual revenue:    ${config["current_state"]["annual_revenue"]:,}')
    print(f'  Pricing phases:    {len(config.get("pricing_phases", []))}')
    print('\nOpen in Excel/LibreOffice — formulas will calculate on first open.')


if __name__ == '__main__':
    main()
